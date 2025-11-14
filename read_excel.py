"""
Script to read Excel file and extract images
"""
import sys
import os

try:
    import openpyxl
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

def read_excel_with_openpyxl(file_path):
    """Read Excel file using openpyxl (supports images)"""
    print(f"Reading Excel file: {file_path}\n")
    
    wb = load_workbook(file_path, data_only=True)
    
    print(f"Sheet names: {wb.sheetnames}\n")
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n{'='*60}")
        print(f"Sheet: {sheet_name}")
        print(f"{'='*60}\n")
        
        # Read cell data
        print("Cell Data:")
        print("-" * 60)
        for row in sheet.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                print(row)
        
        # Check for images
        if hasattr(sheet, '_images') and sheet._images:
            print(f"\nFound {len(sheet._images)} image(s) in this sheet")
            for idx, img in enumerate(sheet._images):
                print(f"  Image {idx + 1}: Found")
                try:
                    if hasattr(img, 'anchor'):
                        print(f"    Location: Cell area")
                except:
                    pass
        else:
            print("\nNo images found in this sheet")
    
    return wb

def read_excel_with_pandas(file_path):
    """Read Excel file using pandas (faster, but doesn't extract images)"""
    print(f"Reading Excel file: {file_path}\n")
    
    excel_file = pd.ExcelFile(file_path)
    print(f"Sheet names: {excel_file.sheet_names}\n")
    
    for sheet_name in excel_file.sheet_names:
        print(f"\n{'='*60}")
        print(f"Sheet: {sheet_name}")
        print(f"{'='*60}\n")
        
        df = pd.read_excel(excel_file, sheet_name=sheet_name)
        print(df.to_string())
        print()
    
    return excel_file

def extract_images_from_excel(file_path, output_dir="extracted_images"):
    """Extract images from Excel file"""
    if not HAS_OPENPYXL:
        print("openpyxl is required for image extraction. Install with: pip install openpyxl pillow")
        return
    
    try:
        from PIL import Image
        import io
    except ImportError:
        print("Pillow is required for image extraction. Install with: pip install pillow")
        return
    
    os.makedirs(output_dir, exist_ok=True)
    
    wb = load_workbook(file_path)
    
    image_count = 0
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        if hasattr(sheet, '_images') and sheet._images:
            for idx, img in enumerate(sheet._images):
                try:
                    # Get image data
                    image_data = img._data()
                    
                    # Save image
                    image_filename = f"{sheet_name}_image_{idx + 1}.png"
                    image_path = os.path.join(output_dir, image_filename)
                    
                    with open(image_path, 'wb') as f:
                        f.write(image_data)
                    
                    print(f"Extracted: {image_path}")
                    image_count += 1
                except Exception as e:
                    print(f"Error extracting image {idx + 1} from sheet {sheet_name}: {e}")
    
    if image_count == 0:
        print("No images found to extract")
    else:
        print(f"\nExtracted {image_count} image(s) to '{output_dir}' directory")

if __name__ == "__main__":
    excel_file = "Copy of Website Changes and Updates.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"Error: File '{excel_file}' not found")
        sys.exit(1)
    
    print("=" * 60)
    print("Excel File Reader and Image Extractor")
    print("=" * 60)
    print()
    
    # Try to read with openpyxl first (supports images)
    if HAS_OPENPYXL:
        print("Using openpyxl to read Excel file...")
        print()
        wb = read_excel_with_openpyxl(excel_file)
        
        # Try to extract images
        print("\n" + "=" * 60)
        print("Attempting to extract images...")
        print("=" * 60)
        extract_images_from_excel(excel_file)
    elif HAS_PANDAS:
        print("Using pandas to read Excel file (images not supported)...")
        print()
        read_excel_with_pandas(excel_file)
        print("\nNote: To extract images, install openpyxl: pip install openpyxl pillow")
    else:
        print("Error: Neither openpyxl nor pandas is installed.")
        print("Install with: pip install openpyxl pandas pillow")
        sys.exit(1)

